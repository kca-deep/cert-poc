"""
export.py — 검수 '확인(confirmed)' 항목을 Excel/PDF 검증결과로 직렬화.

데이터는 db.list_found_with_review() 한 곳에서만 읽고(단일 소스), 출력 범위 결정
(전부 확인 시 확인만 / 그 외 전체)과 표현만 담당한다. 한글 PDF 는 reportlab 내장
CID 폰트(HYSMyeongJo-Medium, 외부 폰트파일
불필요)를 써서 폐쇄망에서도 동작한다. openpyxl/reportlab 은 무거우므로 함수 내부에서
지연 import 한다(미설치 시 라우터가 잡아 안내).

파일명 규칙: ``{원본문서명}_{timestamp}_검증결과.{xlsx|pdf}``
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from . import db

# holistic findings 는 errorType/location/confidence 가 이미 한글이라 라벨 매핑이
# 필요 없다 (LLM 출력 그대로). 검수상태만 review_actions.action → 한글로 변환.
REVIEW_STATUS_LABELS: dict[str | None, str] = {
    "confirmed": "확인", "rejected": "반려", "pending": "보류", None: "미검수",
}

# 컬럼 정의 (Excel/PDF 공통). (헤더, 너비비율).
_COLUMNS: list[tuple[str, float]] = [
    ("문항", 0.45), ("유형", 0.9), ("위치", 0.9),
    ("인용", 1.6), ("사유", 1.8), ("제안", 1.4),
    ("신뢰도", 0.6), ("검수상태", 0.7), ("검수 코멘트", 1.6),
]


def _select(session_id: str) -> tuple[list[dict], bool]:
    """
    export 대상 항목과 모드를 결정한다.

    - 탐지 항목이 '전부 확인(confirmed)' → 확인 항목만 반환 (confirmed_only=True).
    - 확인이 0건이거나 일부만 확인 → 전체 탐지 항목 반환 (confirmed_only=False).
    어느 쪽이든 행에는 검수상태가 함께 실린다.
    """
    items = db.list_found_with_review(session_id)
    confirmed = [f for f in items if f["action"] == "confirmed"]
    confirmed_only = bool(items) and len(confirmed) == len(items)
    return (confirmed if confirmed_only else items), confirmed_only


def _rows(session_id: str) -> list[list[str]]:
    """대상 finding → 표 행(문자열 리스트). finding 1건 = 행 1개."""
    selected, _ = _select(session_id)
    out: list[list[str]] = []
    for f in selected:
        status = REVIEW_STATUS_LABELS.get(f.get("action"), "미검수")
        out.append([
            str(f["qNumber"]),
            f.get("errorType", "") or "",
            f.get("location", "") or "",
            f.get("quote", "") or "",
            f.get("reason", "") or "",
            f.get("suggestion", "") or "",
            f.get("confidence", "") or "",
            status,
            f.get("comment") or "",
        ])
    return out


def _summary(session_id: str) -> str:
    """PDF 부제용 요약: 모드 + 상태별 건수."""
    items = db.list_found_with_review(session_id)
    counts: dict[str, int] = {}
    for f in items:
        k = REVIEW_STATUS_LABELS.get(f.get("action"), "미검수")
        counts[k] = counts.get(k, 0) + 1
    _, confirmed_only = _select(session_id)
    scope = "확인 항목" if confirmed_only else "전체 탐지 항목"
    parts = [f"{k} {counts[k]}" for k in ("확인", "반려", "보류", "미검수") if counts.get(k)]
    breakdown = " · ".join(parts) if parts else "0"
    return f"{scope} (탐지 {len(items)}건 — {breakdown})"


def export_filename(session_id: str, ext: str, timestamp: str) -> str:
    """``{원본문서명}_{timestamp}_검증결과.{ext}`` (확장자/경로 제거한 stem 사용)."""
    raw = db.get_original_filename(session_id) or session_id
    stem = Path(raw).stem or raw
    return f"{stem}_{timestamp}_검증결과.{ext}"


def has_findings(session_id: str) -> bool:
    """내보낼 탐지 항목(found=1)이 하나라도 있는지."""
    return bool(db.list_found_with_review(session_id))


# ── Excel ────────────────────────────────────────────────────────
def build_xlsx(session_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "검증결과"

    headers = [c[0] for c in _COLUMNS]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = _rows(session_id)
    wrap = Alignment(vertical="top", wrap_text=True)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    # 컬럼 너비 — 비율 × 기준치.
    for col, (_, ratio) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(8, ratio * 18)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────
_PDF_FONT = "HYSMyeongJo-Medium"  # reportlab 번들 CID 한글폰트 (외부 파일 불필요)


def _ensure_pdf_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    if _PDF_FONT not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont(_PDF_FONT))
    return _PDF_FONT


def build_pdf(session_id: str, title: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    font = _ensure_pdf_font()
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell", parent=styles["Normal"], fontName=font, fontSize=8, leading=11,
    )
    head_style = ParagraphStyle(
        "head", parent=cell_style, textColor=colors.white, fontSize=8.5,
    )
    title_style = ParagraphStyle(
        "title", parent=styles["Title"], fontName=font, fontSize=15,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )

    rows = _rows(session_id)
    page_w = landscape(A4)[0] - 24 * mm
    ratios = [c[1] for c in _COLUMNS]
    total = sum(ratios)
    col_widths = [page_w * r / total for r in ratios]

    data: list[list[Any]] = [[Paragraph(c[0], head_style) for c in _COLUMNS]]
    for r in rows:
        data.append([Paragraph(_esc(v), cell_style) for v in r])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9CA3AF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F3F4F6")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    story = [
        Paragraph(_esc(title), title_style),
        Spacer(1, 4 * mm),
        Paragraph(_esc(_summary(session_id)), cell_style),
        Spacer(1, 3 * mm),
        table,
    ]
    doc.build(story)
    return buf.getvalue()


def _esc(s: str) -> str:
    """Paragraph 는 미니 HTML 을 해석 → &/</> 이스케이프."""
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
